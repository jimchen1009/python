import os
import re
import shutil

import pandas as pd
from openpyxl import load_workbook
from openpyxl.chart import PieChart, Reference, BarChart, LineChart
from openpyxl.chart.label import DataLabelList
from openpyxl.styles import numbers
from openpyxl.utils import get_column_letter

# 文件路径配置
input_csv_path = 'E:/chrome/每分钟事件次数-data-as-joinbyfield-2025-03-13 11_24_12.csv'  # 协议统计原始数据
input_txt_path = 'C:/Users/chenjingjun/Desktop/协议说明.txt'  # 协议业务说明
output_path = 'C:/Users/chenjingjun/Desktop/协议分析报告.xlsx'


def load_data():
    # 目标文件路径配置
    # 获取协议说明文件的目录
    destination_directory = os.path.dirname(input_txt_path)
    # 新文件名
    new_file_name = '原数据.csv'
    # 目标完整路径
    destination_file_path = os.path.join(destination_directory, new_file_name)

    # 复制文件并重命名
    shutil.copy(input_csv_path, destination_file_path)

    """加载数据源"""
    # 从CSV加载原始数据（假设包含Time列和其他协议列）
    raw_df = pd.read_csv(
        destination_file_path,
        parse_dates=['Time'],  # 自动解析时间列
        thousands=',',  # 处理千分位分隔符
        encoding='utf-8',  # 根据实际编码调整
        skiprows=1  # 跳过第一行
    )

    # # 从TXT加载协议说明（格式：协议,业务说明）
    # tasks_df = pd.read_csv(
    #     input_txt_path,
    #     header=None,
    #     names=['协议', '业务说明'],
    #     sep='\s*,\s*',  # 匹配可能带空格的逗号
    #     engine='python',
    #     encoding='utf-8',
    #     na_filter=False  # 允许读取空值
    # )

    # 读取文件内容
    with open(input_txt_path, 'r', encoding='utf-8') as file:
        lines = file.readlines()
    # 使用正则表达式解析每一行
    data = []
    for line in lines:
        # 使用正则表达式匹配协议和业务说明
        match = re.match(r'([^,]+),\s*(.*)', line.strip())
        if match:
            protocol = match.group(1).strip()
            business_description = match.group(2).strip()
            data.append({'协议': protocol, '业务说明': business_description})
    # 将数据转换为DataFrame
    tasks_df = pd.DataFrame(data)

    return raw_df, tasks_df


def process_data(df, tasks_df):
    """通用数据处理函数"""
    # 保留原始数据副本
    temp_df = df.copy()

    # 移除时间列（保留到分组后再删除）
    if 'Time' in temp_df.columns:
        temp_df = temp_df.drop('Time', axis=1)

    # 转换数值类型
    for col in temp_df.columns:
        temp_df[col] = pd.to_numeric(temp_df[col], errors='coerce').fillna(0)

    # 计算总数
    totals = temp_df.sum()
    totals_df = pd.DataFrame(totals, columns=['总数']).reset_index()
    totals_df.rename(columns={'index': '协议'}, inplace=True)

    # 合并业务说明
    merged_df = pd.merge(totals_df, tasks_df, on='协议', how='outer')
    merged_df['总数'].fillna(0, inplace=True)

    # 拆分协议字段
    merged_df[['ext', 'cmd']] = merged_df['协议'].str.split('/', expand=True)

    #
    merged_df['ext'] = pd.to_numeric(merged_df['ext'], errors='coerce')
    merged_df['cmd'] = pd.to_numeric(merged_df['cmd'], errors='coerce')

    # 排序并整理列顺序
    return merged_df.sort_values(by=['ext', 'cmd'])[['协议', '总数', '业务说明']]


def main_excel_formulas(ws):
    # 获取数据范围
    max_row = ws.max_row
    max_column = ws.max_column
    # 在最后一行添加求和公式
    for col in range(2, max_column + 1):  # 从第二列开始到最后一列
        # 使用 Excel 的 SUM 函数，动态计算每列的和
        ws.cell(row=max_row + 1, column=col,
                value=f'=SUM({ws.cell(row=2, column=col).coordinate}:{ws.cell(row=max_row, column=col).coordinate})')

    # 在第一列的最后一行添加总和标签
    ws.cell(row=max_row + 1, column=1, value='总和')  # 在第一列添加标签
    #
    ws.cell(row=max_row + 1, column=max_column + 1,
            value=f'=SUM({ws.cell(row=max_row + 1, column=1).coordinate}:{ws.cell(row=max_row + 1, column=max_column).coordinate})')

    # 在求和结果下方添加百分比公式
    for col in range(2, max_column + 1):  # 从第二列开始到最后一列
        # 使用 Excel 的公式计算百分比
        ws.cell(row=max_row + 2, column=col,
                value=f'={ws.cell(row=max_row + 1, column=col).coordinate}/{ws.cell(row=max_row + 1, column=max_column + 1).coordinate}')
        # 设置格式为百分比
        ws.cell(row=max_row + 2, column=col).number_format = '0.0000%'

    # 设置自动列宽
    col = next(ws.iter_cols(min_col=1, max_col=1))
    max_length = max(len(str(cell.value)) for cell in col)
    ws.column_dimensions[get_column_letter(col[0].column)].width = max_length + 2

    # 绘制曲线图
    line_chart = LineChart()
    line_chart.title = "协议走势图"
    line_chart.x_axis.title = "时间"
    line_chart.y_axis.title = "数值"

    # 设置数据范围
    data = Reference(ws, min_col=2, min_row=1, max_col=max_column, max_row=max_row)
    line_chart.add_data(data, titles_from_data=True)

    # 设置分类轴（时间轴）， 排除标题行
    labels = Reference(ws, min_col=1, min_row=2, max_row=max_row)
    line_chart.set_categories(labels)

    # 设置曲线颜色和标签
    color_cycle = [
        'FF0000',  # 红
        '00B050',  # 绿
        '0070C0',  # 蓝
        'FFC000',  # 黄
        'FF00FF',  # 品红
        '00FFFF',  # 青色
        '800080',  # 紫色
        'FFA500',  # 橙色
        'FFD700',  # 金色
        'C0C0C0',  # 银色
        '808080',  # 灰色
        '000000',  # 黑色
        'FF6347',  # 番茄色
        '4682B4',  # 钢蓝色
        'ADFF2F',  # 黄绿色
        'FF1493',  # 深粉色
        '1E90FF',  # 道奇蓝
        '32CD32',  # 莱姆绿
        'FF4500',  # 橙红色
        'DA70D6',  # orchid
        '7FFF00',  # 查特酒绿
        'D2691E',  # 巧克力色
        'FF8C00',  # 深橙色
        'FFDAB9',  # 桃色
        'FFB6C1',  # 浅粉色
        'B22222',  # 火砖色
        '5F9EA0',  # 深海洋蓝
        'FF7F50',  # 珊瑚色
        '6495ED',  # 矢车菊蓝
        'FFFACD',  # 柠檬奶油色
        'FFE4B5',  # 桃色
        'F0E68C',  # 卡其色
    ]
    color_len = len(color_cycle)
    for i, series in enumerate(line_chart.series):
        # 设置曲线颜色（循环使用预设颜色）
        series.graphicalProperties.line.solidFill = color_cycle[i % color_len]

        # 显示数据
        series.marker.symbol = None  # 不显示数据点标记
        series.marker.graphicalProperties.solidFill = color_cycle[i % color_len]
        series.marker.graphicalProperties.line.solidFill = 'FFFFFF'

        # 显示数据标签
        series.dLbls = DataLabelList()
        series.dLbls.showVal = False  # 不显示数值
        series.dLbls.showSerName = False  # 不在数据点上显示序列名称

    line_chart.height = 20
    line_chart.width = 50

    # 添加图例到图表
    line_chart.legend.position = 'r'  # 使用有效的图例位置

    # 将曲线图添加到工作表的最后一行下方的第一列
    ws.add_chart(line_chart, f"B{max_row + 4}")  # 在最后一行下方的第一列放置图表


def add_excel_formulas(ws):
    """添加Excel公式和格式"""
    # 插入百分比列
    ws.insert_cols(3)
    ws['C1'] = '百分比'

    # 获取数据行数，列数
    max_row = ws.max_row

    # 添加百分比公式
    for row in range(2, max_row + 1):
        ws[f'C{row}'] = f'=B{row}/$B${max_row + 1}'
        ws[f'C{row}'].number_format = numbers.FORMAT_PERCENTAGE_00

    # 添加总和行
    ws[f'A{max_row + 1}'] = '总计'
    ws[f'B{max_row + 1}'] = f'=SUM(B2:B{max_row})'

    # 设置自动列宽
    for col in ws.columns:
        max_length = max(len(str(cell.value)) for cell in col)
        ws.column_dimensions[get_column_letter(col[0].column)].width = max_length + 2

    # 添加图表（仅在有效数据存在时）
    if max_row > 2:  # 标题行+至少1行数据
        max_data_row = max_row - 1  # 排除总计行

        # === 柱形图 ===
        bar = BarChart()
        bar.title = "协议数量（总数）"
        # 数据范围：A列协议，B列总数
        data = Reference(ws, min_col=2, min_row=2, max_row=max_data_row)
        bar.add_data(data, titles_from_data=True)
        # 数据范围：A列协议，B列总数
        labels = Reference(ws, min_col=1, min_row=2, max_row=max_data_row)
        bar.set_categories(labels)
        bar.varyColors = False  # 使用单一颜色
        bar.series[0].graphicalProperties.solidFill = "4472C4"  # 标准蓝色
        # 设置坐标轴标题
        bar.y_axis.title = '数量'
        bar.x_axis.title = '协议类型'
        # 调整样式
        bar.height = 14
        bar.width = 20 + max_row * 0.8
        bar.style = 13  # 使用预定义样式
        # 数据标签配置
        bar.dataLabels = DataLabelList()
        bar.y_axis.majorGridlines = None  # 移除网格线
        bar.dataLabels.position = 'outEnd'  # 标签显示在柱形末端
        # 放置
        ws.add_chart(bar, "F2")  # 放置在图表的G2位置

        # === 饼图 ===
        pie = PieChart()
        pie.title = "协议分布（百分比）"

        data = Reference(ws, min_col=3, min_row=2, max_row=max_data_row)
        pie.add_data(data, titles_from_data=True)
        # 数据范围：A列协议，C列百分比
        labels = Reference(ws, min_col=1, min_row=3, max_row=max_data_row)
        pie.set_categories(labels)
        # 设置数据标签显示百分比
        pie.dataLabels = DataLabelList()
        pie.dataLabels.showVal = True
        # 调整图表位置和大小
        pie.height = 25  # 高度14厘米
        pie.width = 30  # 宽度18厘米

        ws.add_chart(pie, "F32")  # 放置在图表的G2位置


def main():
    # 加载数据
    raw_df, tasks_df = load_data()

    # 验证任务表格式
    if tasks_df.shape[1] != 2:
        raise ValueError("任务表需要包含两列：协议和业务说明")
    tasks_df.columns = ['协议', '业务说明']

    # 处理时间格式
    raw_df['Time'] = pd.to_datetime(raw_df['Time'],
                                    format='%m/%d/%Y, %I:%M:%S %p',
                                    errors='coerce')

    # 过滤无效时间
    raw_df = raw_df.dropna(subset=['Time'])

    # 创建日期列
    raw_df['Date'] = raw_df['Time'].dt.date
    # 创建小时列
    raw_df['Hour'] = raw_df['Time'].dt.hour

    # 创建Excel写入对象
    with pd.ExcelWriter(output_path, engine='openpyxl') as writer:

        # 处理总统计
        total_df = process_data(raw_df.drop(columns=['Date', 'Hour']), tasks_df)
        total_df.to_excel(writer, sheet_name='总统计', index=False)
        # 处理源数据
        raw_df.drop(columns=['Date', 'Hour']).to_excel(writer, sheet_name='源数据', index=False)

        # 按日期分组处理
        for date, group in raw_df.groupby('Date'):
            # 生成日期字符串（格式：YYYY年MM月DD日）
            date_str = date.strftime("%Y年%m月%d日")

            # 处理当日数据
            daily_df = process_data(group.drop(columns=['Date', 'Hour']), tasks_df)

            # 写入工作表（名称示例：2024年12月06）
            daily_df.to_excel(writer, sheet_name=date_str, index=False)

        # # 按日期分组处理
        # for date, group in raw_df.groupby('Date'):
        #     # 生成日期字符串（格式：YYYY年MM月DD日）
        #     day_str = date.strftime("%d日")
        #     # 按小时分组处理
        #     for hour, hour_group in group.groupby('Hour'):
        #         # 生成小时字符串（格式：YYYY年MM月DD日HH时）
        #         hour_str = f"{day_str}{hour:02d}时"
        #
        #         # 处理当小时数据
        #         hourly_df = process_data(hour_group.drop(columns=['Date', 'Hour']), tasks_df)
        #
        #         # 写入工作表（名称示例：2024年12月06日00时）
        #         hourly_df.to_excel(writer, sheet_name=hour_str, index=False)

    # 添加公式和格式
    wb = load_workbook(output_path)
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        if sheet_name == "源数据":
            main_excel_formulas(ws)
        else:
            add_excel_formulas(ws)
    wb.save(output_path)

    print(f"处理完成，结果已保存至：{output_path}")


if __name__ == "__main__":
    main()
